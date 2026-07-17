from datetime import datetime

import pytest
from openpyxl import Workbook

from src.register_invoice import (
    format_invoice_no,
    is_duplicate_invoice_no,
    read_invoice_from_excel,
    register_invoice,
    get_invoice_file,
    parse_args,
)


class MockSheet:

    def col_values(self, column):
        return [
            "請求書No",
            "2607-01",
            "2607-02",
        ]

class MockRegisterSheet:

    def __init__(self):
        self.rows = []

    def col_values(self, column):
        return [
            "請求書No",
        ]

    def append_row(self, row):
        self.rows.append(row)


class MockDuplicateSheet:

    def __init__(self):
        self.rows = []

    def col_values(self, column):
        return [
            "請求書No",
            "2607-01",
        ]

    def append_row(self, row):
        self.rows.append(row)


class MockErrorSheet:

    def col_values(self, column):
        return [
            "請求書No",
        ]

    def append_row(self, row):
        raise Exception("登録エラー")
    

def test_is_duplicate_invoice_no_true():

    sheet = MockSheet()

    result = is_duplicate_invoice_no(
        sheet,
        "2607-01"
    )

    assert result is True


def test_is_duplicate_invoice_no_false():

    sheet = MockSheet()

    result = is_duplicate_invoice_no(
        sheet,
        "2607-03"
    )

    assert result is False


def test_format_invoice_no_datetime():

    result = format_invoice_no(
        datetime(2026, 7, 15)
    )

    assert result == "2607-15"


def test_format_invoice_no_string():

    result = format_invoice_no(
        "2607-01"
    )

    assert result == "2607-01"


def test_read_invoice_from_excel(tmp_path):

    # テスト用Excel作成
    file_path = tmp_path / "test_invoice.xlsx"

    wb = Workbook()

    ws = wb.active
    ws.title = "請求書"

    ws["F2"] = "2607-01"
    ws["F3"] = datetime(2026, 7, 15)
    ws["F4"] = datetime(2026, 8, 31)
    ws["B3"] = "ミヤシタ技研"
    ws["B4"] = "HP更新"
    ws["F31"] = 50000

    wb.save(file_path)


    # 関数実行
    invoice = read_invoice_from_excel(file_path)


    # 結果確認
    assert invoice["請求書No"] == "2607-01"
    assert invoice["取引先"] == "ミヤシタ技研"
    assert invoice["案件名"] == "HP更新"
    assert invoice["金額"] == 50000


def test_read_invoice_from_excel_file_not_found(tmp_path):

    file_path = tmp_path / "not_found.xlsx"

    invoice = read_invoice_from_excel(file_path)

    assert invoice is None


def test_read_invoice_from_excel_sheet_not_found(tmp_path):

    file_path = tmp_path / "no_sheet.xlsx"

    wb = Workbook()

    ws = wb.active
    ws.title = "Sheet1"

    wb.save(file_path)

    invoice = read_invoice_from_excel(file_path)

    assert invoice is None


def test_register_invoice_success():

    sheet = MockRegisterSheet()

    invoice = {
        "請求書No": "2607-01",
        "送付日": "2026/07/15",
        "支払期限": "2026/08/31",
        "取引先": "ミヤシタ技研",
        "案件名": "HP更新",
        "金額": 50000,
        "入金日": "",
    }

    result = register_invoice(
        sheet,
        invoice
    )

    assert result is True
    assert len(sheet.rows) == 1
    assert sheet.rows[0][0] == "2607-01"


def test_register_invoice_duplicate():

    sheet = MockDuplicateSheet()

    invoice = {
        "請求書No": "2607-01",
        "送付日": "2026/07/15",
        "支払期限": "2026/08/31",
        "取引先": "ミヤシタ技研",
        "案件名": "HP更新",
        "金額": 50000,
        "入金日": "",
    }

    result = register_invoice(
        sheet,
        invoice
    )

    assert result is False
    assert len(sheet.rows) == 0


def test_register_invoice_error():

    sheet = MockErrorSheet()

    invoice = {
        "請求書No": "2607-01",
        "送付日": "2026/07/15",
        "支払期限": "2026/08/31",
        "取引先": "ミヤシタ技研",
        "案件名": "HP更新",
        "金額": 50000,
        "入金日": "",
    }

    result = register_invoice(
        sheet,
        invoice
    )

    assert result is False


def test_get_invoice_file_success(tmp_path):

    file_path = tmp_path / "2607-01_test.xlsx"

    file_path.touch()

    result = get_invoice_file(file_path)

    assert result == file_path


def test_parse_args_no_argument(monkeypatch):

    monkeypatch.setattr(
        "sys.argv",
        ["register_invoice.py"]
    )

    with pytest.raises(SystemExit):
        parse_args()


def test_get_invoice_file_not_found(tmp_path):

    file_path = tmp_path / "not_found.xlsx"

    with pytest.raises(FileNotFoundError):
        get_invoice_file(file_path)


def test_register_invoice_no_empty():

    sheet = MockRegisterSheet()

    invoice = {
        "請求書No": "",
        "送付日": "",
        "支払期限": "",
        "取引先": "",
        "案件名": "",
        "金額": 0,
        "入金日": "",
    }

    result = register_invoice(sheet, invoice)

    assert result is False
    