from pathlib import Path
from datetime import datetime
import logging
from configparser import ConfigParser

import gspread
from gspread.exceptions import SpreadsheetNotFound
from google.oauth2.service_account import Credentials
from openpyxl import load_workbook

# =========================
# 設定
# =========================

# Google Sheets APIのスコープ
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets"
]

# プロジェクトフォルダ
BASE_DIR = Path(__file__).resolve().parent.parent

# 設定ファイル
CONFIG_PATH = BASE_DIR / "config" / "config.ini"


def load_config():
    """設定ファイル読み込み"""

    config = ConfigParser()

    if not CONFIG_PATH.exists():
        raise FileNotFoundError(
            "設定ファイルが見つかりません"
        )

    config.read(CONFIG_PATH, encoding="utf-8")

    return config


def get_settings():
    """設定値取得"""

    config = load_config()

    spreadsheet_id = config["google"]["spreadsheet_id"]
    invoice_dir = BASE_DIR / config["path"]["invoice_dir"]

    return spreadsheet_id, invoice_dir

# ログフォルダ
LOG_DIR = BASE_DIR / "logs"

# ログフォルダ作成
LOG_DIR.mkdir(exist_ok=True)

# ログ設定
logging.basicConfig(
    filename=LOG_DIR / "invoice_manager.log",
    level=logging.INFO,
    encoding="utf-8",
    format="%(asctime)s %(levelname)s %(message)s"
)

logger = logging.getLogger(__name__)


def authenticate(spreadsheet_id):
    """Google Sheetsへ接続"""

    # 認証ファイル
    credentials_path = BASE_DIR / "credentials.json"

    # 認証
    try:
        # 認証ファイル読み込み
        creds = Credentials.from_service_account_file(
            credentials_path,
            scopes=SCOPES
        )

        # Google Sheets API認証
        gc = gspread.authorize(creds)

        # スプレッドシートを開く
        spreadsheet = gc.open_by_key(spreadsheet_id)

        logger.info("Google Sheetsへ接続できました！")

        return spreadsheet

    except FileNotFoundError:
        logger.error("認証ファイルが見つかりません")
        return None

    except SpreadsheetNotFound:
        logger.error("指定したスプレッドシートが見つかりません")
        return None

    except Exception as e:
        logger.exception(f"Google Sheets接続エラー: {e}")
        return None


def read_invoice_from_excel(file_path):
    """
    Excel請求書から請求データを読み込む。

    Args:
        file_path (str): 請求書ファイルのパス

    Returns:
        dict: 請求データ
    """

    try:
        # Excelファイルを読み込み
        wb = load_workbook(file_path, data_only=True)

        # 請求書シートを取得
        ws = wb["請求書"]

        invoice = {
            "請求書No": format_invoice_no(ws["F2"].value),
            "送付日": ws["F3"].value.strftime("%Y/%m/%d"),
            "支払期限": ws["F4"].value.strftime("%Y/%m/%d"),
            "取引先": ws["B3"].value,
            "案件名": ws["B4"].value,
            "金額": int(ws["F31"].value),
            "入金日": ""
        }

        return invoice

    except FileNotFoundError:
        logger.error("請求書ファイルが見つかりません")
        return None

    except KeyError:
        logger.error("請求書シートが見つかりません")
        return None

    except Exception as e:
        logger.exception(f"Excel読み込みエラー: {e}")
        return None

def format_invoice_no(value):
    """請求書Noを文字列形式へ変換"""

    if isinstance(value, datetime):
        return value.strftime("%y%m-%d")

    return str(value)

def is_duplicate_invoice_no(sheet, invoice_no):
    """請求書Noの重複チェック"""
    
    # 請求書No列を取得（1行目の見出しを除外）
    invoice_nos = sheet.col_values(1)[1:]

    # 文字列として比較し、既存データに存在する場合はTrueを返す
    return str(invoice_no) in invoice_nos   


def register_invoice(sheet, invoice):
    """Googleスプレッドシートへ登録"""
  
    invoice_no = invoice["請求書No"]

    if is_duplicate_invoice_no(sheet, invoice_no):
        logger.warning(f"登録済みの請求書Noです: {invoice_no}")
        return False

    row = [
        invoice["請求書No"],
        invoice["送付日"],
        invoice["支払期限"],
        invoice["取引先"],
        invoice["案件名"],
        invoice["金額"],
        invoice["入金日"],
    ]

    try:
        sheet.append_row(row)

    except Exception as e:
        logger.exception(f"登録エラー: {e}")
        return False

    return True

def main():
    try:
        spreadsheet_id, invoice_dir = get_settings()

    except Exception as e:
        logger.error(f"設定読み込みエラー: {e}")
        return

    spreadsheet = authenticate(spreadsheet_id)

    if spreadsheet is None:
        return

    sheet = spreadsheet.sheet1

    file_path = (
        invoice_dir
        / "2607-02_ミヤシタ技研_岡山営業所関連_HP更新・新ロゴ.xlsx"
    )

    invoice = read_invoice_from_excel(file_path)

    if invoice is None:
        return

    result = register_invoice(sheet, invoice)

    if result:
        logger.info("請求データを登録しました！")
    else:
        logger.info(f"請求データの登録を中止しました。請求書No:{invoice['請求書No']}")

if __name__ == "__main__":
    main()