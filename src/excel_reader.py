from datetime import datetime

from openpyxl import load_workbook

from src.register_invoice import logger


def format_date(value):
    """日付を文字列へ変換"""

    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.strftime("%Y/%m/%d")

    return str(value)


def format_amount(value):
    """金額を整数へ変換"""

    if value is None:
        return 0

    return int(value)


def format_text(value):
    """文字列項目を変換"""

    if value is None:
        return ""

    return str(value)


def format_invoice_no(value):
    """請求書Noを文字列形式へ変換"""

    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.strftime("%y%m-%d")

    return str(value)


def read_invoice_from_excel(file_path):
    """Excel請求書から請求データを読み込む"""

    wb = None

    try:

        load_options = {
            "data_only": True
        }

        if file_path.suffix.lower() == ".xlsm":
            load_options["keep_vba"] = True

        wb = load_workbook(
            file_path,
            **load_options
        )

        ws = wb["請求書"]

        invoice = {
            "請求書No": format_invoice_no(ws["F2"].value),
            "送付日": format_date(ws["F3"].value),
            "支払期限": format_date(ws["F4"].value),
            "取引先": format_text(ws["B3"].value),
            "案件名": format_text(ws["B4"].value),
            "金額": format_amount(ws["F31"].value),
            "入金日": "",
        }

        return invoice

    except FileNotFoundError:
        logger.error("請求書ファイルが見つかりません")
        return None

    except KeyError:
        logger.error("請求書シートが見つかりません")
        return None

    except Exception as e:
        logger.exception(f"Excel読み込みエラー: {e}")
        return None

    finally:
        if wb:
            wb.close()